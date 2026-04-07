import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import Workbook

# Dados para a planilha
data = {
    'Data': ['31/03/2026', '31/03/2026', '01/04/2026', '01/04/2026'],
    'Cliente': ['Carlos Silva', 'Maria Oliveira', 'João Pereira', 'Ana Costa'],
    'Produto/Serviço': ['Consultoria IA', 'Licença Software', 'Mentoría VIP', 'E-book Estratégias'],
    'Qtd': [1, 2, 1, 5],
    'Valor Unitário': [1500.00, 250.00, 3000.00, 49.90],
    'Valor Total': [1500.00, 500.00, 3000.00, 249.50],
    'Canal': ['WhatsApp', 'Instagram', 'LinkedIn', 'Tráfego Pago'],
    'Status': ['Pago', 'Pendente', 'Pago', 'Pago']
}

df = pd.DataFrame(data)
target_file = r'G:\Meu Drive\controle_vendas_premium.xlsx'

# Usando o ExcelWriter para aplicar estilos
with pd.ExcelWriter(target_file, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Vendas')
    worksheet = writer.sheets['Vendas']
    
    # Estilizando o Cabeçalho (Azul Marinho com Branco)
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        
    # Estilizando o corpo da planilha
    zebra_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    currency_format = '"R$" #,##0.00'
    
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
        # Zebrado (linhas pares com cor de fundo)
        if row_idx % 2 == 0:
            for cell in row:
                cell.fill = zebra_fill
        
        # Formatação de Moeda para as colunas 5 e 6 (Valor Unitário e Valor Total)
        worksheet.cell(row=row_idx, column=5).number_format = currency_format
        worksheet.cell(row=row_idx, column=6).number_format = currency_format
        
        # Alinhamento Central para Qtd e Status
        worksheet.cell(row=row_idx, column=4).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_idx, column=8).alignment = Alignment(horizontal='center')

    # Ajuste automático das larguras das colunas
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 4)
        worksheet.column_dimensions[column_letter].width = adjusted_width

print(f"Sucesso! Planilha salva em: {target_file}")
